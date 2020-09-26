#! python3
# portfolio.py - Pull stock data from yfinance API and update portfolio tracker accordingly
# the script will run once a day after market close

import sys
from datetime import datetime

import openpyxl
import yfinance as yf

portfolio = 'portfolio.xlsx'
portfolio_history = 'portfolio_history.xlsx'

if datetime.today().weekday() in range(5,6):
    print("The market is closed on weekends. Exiting program.")
    sys.exit()

wb = openpyxl.load_workbook(portfolio)
destination = openpyxl.load_workbook(portfolio_history)

current = wb['Current']

# store prior day's market value before updating for current day's prices
current.cell(row=2, column=16).value = current.cell(row=7, column=8).value

# Loop through rows to grab ticker references which will be used to pull stock data from yf
for row_num in range(2, current.max_row): # skipping first row because it's the header. leaving last row because it's the total
    stock_ticker = (current.cell(row=row_num, column=2)).value
    print(f"Gathering {stock_ticker} data.")
    stock_ticker_string = str(stock_ticker)
    stock_info = yf.Ticker(stock_ticker_string)
    stock_data = stock_info.history(period = '5d')
    current_quote = (stock_data.tail(1)['Close'].iloc[0])
    yesterday_quote = (stock_data.tail(2)['Close'].iloc[0])
    day_high = (stock_data.tail(1))['High'].iloc[0]
    day_low = (stock_data.tail(1))['Low'].iloc[0]
    day_dividends = (stock_data.tail(1))['Dividends'].iloc[0]

    current.cell(row=row_num, column=7).value = current_quote
    current.cell(row=row_num, column=3).value = (current_quote - yesterday_quote)/yesterday_quote
    current.cell(row=row_num, column=8).value = current.cell(row=row_num, column=7).value * current.cell(row=row_num, column = 4).value
    current.cell(row=row_num, column=9).value = (current.cell(row=row_num, column=8).value - current.cell(row=row_num, column=6).value)
    current.cell(row=row_num, column=10).value = current.cell(row=row_num, column=9).value / current.cell(row=row_num, column=6).value * 100
    current.cell(row=row_num, column=11).value = day_high
    current.cell(row=row_num, column=12).value = day_low
    current.cell(row=row_num, column=13).value = day_dividends * current.cell(row=row_num, column=4).value

    if day_dividends:
        print(f"Recording reinvestment of {stock_ticker} dividend payment.")
        current.cell(row=row_num, column=4).value += current.cell(row=row_num, column=13).value / current_quote
        current.cell(row=row_num, column=6).value += current.cell(row=row_num, column=13).value

print("Calculating today's gain/loss...")
current.cell(row=7, column=8).value = current.cell(row=2, column=8).value + current.cell(row=3, column=8).value + current.cell(row=4, column=8).value + current.cell(row=5, column=8).value + current.cell(row=6, column=8).value
current.cell(row=1, column=16).value = round(float(current.cell(row=7, column=8).value) - float(current.cell(row=2, column=16).value),2)

# create tab in portfolio_history.xlsx file to store the day's portfolio value,
# naming the tabs as the current date
destination.create_sheet(datetime.today().strftime('%y-%m-%d'))
destination.save(portfolio_history)
export_tab = destination[datetime.today().strftime('%y-%m-%d')]

# get data from portfolio.xlsx in order to copy it to portfolio_history.xlsx
mr = current.max_row
mc = current.max_column

for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        c = current.cell(row = i, column = j)
        export_tab.cell(row = i, column = j).value = c.value

print("Data copied to portfolio history file. Saving files...")
wb.save(portfolio)
destination.save(portfolio_history)